import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side



# Take input
input("The last inputs will ask for the ground truth values, have these ready, if you don't CTRL+C to stop. \
\nPress Enter to continue.\n")
path_dir = input("What is the path of the run to compile the data from? (ex: ../runs/val/[NAME]) \n>")
file_name = input("What do you want to name the finished spreadsheet? \n>")
# sheet_mode = input("What class is this for? (0 = berries, 1 = bushes) \n>") # Not configured yet
table_name = input("What should the table be named? Tables should be the name of the Validation Dataset used.\n>")
sheet_name = input("What should the sheet be named? Sheets should be the name of weight used.\n>")
blue_gt = int(input("What was the ground truth for the blue count? (Roboflow ground truth)\n)>"))
green_gt = int(input("What was the ground truth for the green count? (Roboflow ground truth)\n>"))



# Create a workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = sheet_name

# Parsed data
data = []
conf_threshold = -1
iou_threshold = -1

# Used for shifting starting location of tables [# of Columns, # of Rows]
table_size = [10, 3]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

thick_border_r = Border(left=Side(style='thin'),
                         right=Side(style='thick'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

thick_border_b = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thick'))

thick_border_t = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thick'),
                         bottom=Side(style='thin'))

light_grey = PatternFill(start_color='aeaeae',
                           end_color='aeaeae',
                           fill_type='solid')


### START OF BODY ###

def parse_data(path_to_data):
    global data, conf_threshold, iou_threshold

    # Parse the Validation_Metrics file
    file_data = []
    with open(path_to_data, 'r') as file:
        for line in file:
            line = line.strip().split(',')
            file_data.append(line)
    # print(data)

    # Format parsed data
    data = [
        [file_data[0][0], blue_gt + green_gt, 'A', 'A', 'A',
         int(file_data[0][2]), round(float(file_data[0][3]), 4), round(float(file_data[0][4]), 4),
         round(float(file_data[0][5]), 4), round(float(file_data[0][6]), 4)],
        [file_data[1][0], blue_gt, 'B1', 'B2', 'B3',
         int(file_data[1][2]), round(float(file_data[1][3]), 4), round(float(file_data[1][4]), 4),
         round(float(file_data[1][5]), 4), round(float(file_data[1][6]), 4)],
        [file_data[2][0], green_gt, 'C1', 'C2', 'C3',
         int(file_data[2][2]), round(float(file_data[2][3]), 4), round(float(file_data[2][4]), 4),
         round(float(file_data[2][5]), 4), round(float(file_data[2][6]), 4)]
    ]

    conf_threshold = file_data[0][7]
    iou_threshold = file_data[0][8]

def initialize_new_headers():
    bold_font = Font(bold=True)
    dark_grey = PatternFill(start_color='999999',
                           end_color='999999',
                           fill_type='solid')
    medium_grey = PatternFill(start_color='cccccc',
                           end_color='cccccc',
                           fill_type='solid')


    # ws.merge_cells('A1:L1')
    ws['A1'] = table_name
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = bold_font
    ws['A1'].fill = dark_grey
    ws['A1'].border = thin_border

    # ws.merge_cells('C2:L2')
    ws['C2'] = 'IOU Threshold'
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'].font = bold_font
    ws['C2'].fill = medium_grey
    ws['C2'].border = thin_border

    # ws.merge_cells('A5:A7')
    ws['A5'] = 'Confidence Threshold'
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws['A5'].font = bold_font
    ws['A5'].fill = medium_grey
    ws['A5'].border = thin_border

    for row in ws['A2:A4']:
        for cell in row:
            cell.fill = medium_grey
            cell.border = thin_border

    ws['B2'].fill = medium_grey
    ws['B2'].border = thin_border
    ws['B3'].fill = light_grey
    ws['B3'].border = thin_border
    ws['B4'].fill = light_grey
    ws['B4'].border = thin_border

def populate_table():
    global data, conf_threshold, iou_threshold

    table_starting_cell = find_spot_in_table(float(conf_threshold), float(iou_threshold))

    start_col = table_starting_cell[0]
    start_row = table_starting_cell[1]

    iou_starting_letter = get_column_letter(start_col)
    iou_ending_letter = get_column_letter(start_col + table_size[0] - 1)

    # Confidence Threshold Header
    # print(ws[f'B{start_row }'].value)
    if ws[f'B{start_row}'].value is None:
        ws.merge_cells(f'B{start_row}:B{start_row + table_size[1] - 1}')
        ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{start_row}'].value = conf_threshold
        ws[f'B{start_row}'].fill = light_grey
        ws[f'B{start_row}'].border = thick_border_r

    # IOU Threshold Header & Metric Headers
    # print(ws[f'{iou_starting_letter}3'].value)
    if ws[f'{iou_starting_letter}3'].value is None:
        ws.merge_cells(f'{iou_starting_letter}3:{iou_ending_letter}3')
        ws[f'{iou_starting_letter}3'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{iou_starting_letter}3'].value = iou_threshold
        ws[f'{iou_starting_letter}3'].fill = light_grey
        ws[f'{iou_starting_letter}3'].border = thin_border


        # Define the list of values
        headers = ['Class', 'GT', 'TP', 'FP', 'FN', 'Det', 'Pre', 'Rec', 'mAP:0.5', 'mAP:0.5:0.95']

        # Create a list of the column letters
        columns = [get_column_letter(i) for i in range(start_col, start_col + table_size[0])]

        # Iterate over the values and columns and put the values in the specified columns
        for header, column in zip(headers, columns):
            ws[f'{column}4'] = header
            ws[f'{column}4'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{column}4'].fill = light_grey
            ws[f'{column}4'].border = thick_border_b

    for row_index, row in enumerate(data, start=start_row):
        for col_index, value in enumerate(row, start=start_col):
            if value == 'A':
                value = f'=SUM({get_column_letter(col_index)}{row_index + 1}:{get_column_letter(col_index)}{row_index + 2})'
            elif value == 'B1' or value == 'C1':
                value = f'=ROUND({get_column_letter(col_index + 3)}{row_index}*{get_column_letter(col_index + 4)}{row_index}, 0)'
            elif value == 'B2' or value == 'C2':
                value = f'={get_column_letter(col_index + 3)}{row_index}-{get_column_letter(col_index - 1)}{row_index}'
            elif value == 'B3' or value == 'C3':
                value = f'=ROUND(({get_column_letter(col_index - 2)}{row_index}-({get_column_letter(col_index - 2)}{row_index}*{get_column_letter(col_index + 3)}{row_index}))/{get_column_letter(col_index + 3)}{row_index}, 0)'

            ws.cell(row=row_index, column=col_index).value = value
            ws.cell(row=row_index, column=col_index).border = thin_border

def find_spot_in_table(conf_thres, iou_thres):
    # [Column, Row], default is C4
    table_starting_cell = [3, 5]

    # print("Finding where conf of " + str(conf_thres) + " and iou of " + str(iou_thres) + " belongs...")

    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value is not None:
            cell_value = float(str(ws.cell(row=row, column=2).value))
            if cell_value < conf_thres:
                row_val = row + table_size[1]
                # print(f'Table could start at row:  {row_val}')
                table_starting_cell[1] = row_val
            elif cell_value == conf_thres:
                row_val = row
                # print(f'Table would start at row:  {row_val}')
                table_starting_cell[1] = row_val

    for column in range(1, ws.max_column + 1):
        if ws.cell(row=3, column=column).value is not None:
            cell_value = float(str(ws.cell(row=3, column=column).value))
            if cell_value < iou_thres:
                col_val = column + table_size[0]
                # print(f'Table could start at column:  {col_val}')
                table_starting_cell[0] = col_val
            elif cell_value == iou_thres:
                col_val = column
                # print(f'Table would start at column:  {col_val}')
                table_starting_cell[0] = col_val

    return table_starting_cell

def update_headers():
    ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
    ws.merge_cells(f'C2:{get_column_letter(ws.max_column)}2')
    ws.merge_cells(f'A5:A{ws.max_row}')

    for row in ws[f'B5:B{ws.max_row}']:
        for cell in row:
            cell.border = thick_border_r

initialize_new_headers()
for folder in os.listdir(path_dir):
    # print(path_dir + "/" + folder + "/Validation_Metrics.txt")
    parse_data(path_dir + "/" + folder + "/Validation_Metrics.txt")
    populate_table()

update_headers()

# Create a table
table = Table(displayName='Table1', ref='A1:AA999')

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style

### END OF BODY ###

# Add the table to the worksheet
ws.add_table(table)

# Save workbook
wb.save(file_name + '.xlsx')

# Gets rid of an error
table.autoFilter = None

# Save workbook again to update autoFilter setting
wb.save(file_name + '.xlsx')

# Close workbook
wb.close()
